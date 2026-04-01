from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import (
    BalanceSnapshot,
    Currency,
    ExpenseCategory,
    IncomeSource,
    StorageAccount,
    Transaction,
)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")


async def generate_account_export(db: AsyncSession, user_id: int) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    await _sheet_transactions(wb, db, user_id)
    await _sheet_balance_snapshots(wb, db, user_id)
    await _sheet_storage_accounts(wb, db, user_id)
    await _sheet_income_sources(wb, db, user_id)
    await _sheet_expense_categories(wb, db, user_id)
    await _sheet_currencies(wb, db, user_id)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _sheet_transactions(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Transactions")
    _write_header(ws, ["Date", "Type", "Amount", "Currency", "Storage Location", "Income Source", "Description"])

    result = await db.execute(
        select(Transaction)
        .where(Transaction.user_id == user_id)
        .options(
            selectinload(Transaction.currency),
            selectinload(Transaction.storage_account).selectinload(StorageAccount.storage_location),
            selectinload(Transaction.income_source),
        )
        .order_by(Transaction.date.desc())
    )
    for tx in result.scalars():
        location = (
            tx.storage_account.storage_location.name
            if tx.storage_account and tx.storage_account.storage_location
            else ""
        )
        ws.append([
            tx.date.isoformat(),
            tx.type.value,
            float(tx.amount),
            tx.currency.code if tx.currency else "",
            location,
            tx.income_source.name if tx.income_source else "",
            tx.description or "",
        ])


async def _sheet_balance_snapshots(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Balance Snapshots")
    _write_header(ws, ["Date", "Storage Location", "Currency", "Amount"])

    result = await db.execute(
        select(BalanceSnapshot)
        .where(BalanceSnapshot.user_id == user_id)
        .options(
            selectinload(BalanceSnapshot.storage_account).selectinload(StorageAccount.storage_location),
            selectinload(BalanceSnapshot.storage_account).selectinload(StorageAccount.currency),
        )
        .order_by(BalanceSnapshot.date.desc())
    )
    for snap in result.scalars():
        account = snap.storage_account
        location = account.storage_location.name if account and account.storage_location else ""
        currency = account.currency.code if account and account.currency else ""
        ws.append([
            snap.date.isoformat(),
            location,
            currency,
            float(snap.amount),
        ])


async def _sheet_storage_accounts(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Storage Accounts")
    _write_header(ws, ["Location Name", "Currency Code", "Currency Symbol"])

    result = await db.execute(
        select(StorageAccount)
        .where(StorageAccount.user_id == user_id)
        .options(
            selectinload(StorageAccount.storage_location),
            selectinload(StorageAccount.currency),
        )
    )
    for acc in result.scalars():
        ws.append([
            acc.storage_location.name if acc.storage_location else "",
            acc.currency.code if acc.currency else "",
            acc.currency.symbol if acc.currency else "",
        ])


async def _sheet_income_sources(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Income Sources")
    _write_header(ws, ["Name"])

    result = await db.execute(
        select(IncomeSource)
        .where(IncomeSource.user_id == user_id)
        .order_by(IncomeSource.name)
    )
    for src in result.scalars():
        ws.append([src.name])


async def _sheet_expense_categories(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Expense Categories")
    _write_header(ws, ["Name", "Budgeted Amount", "Tags"])

    result = await db.execute(
        select(ExpenseCategory)
        .where(ExpenseCategory.user_id == user_id)
        .order_by(ExpenseCategory.name)
    )
    for cat in result.scalars():
        ws.append([
            cat.name,
            float(cat.budgeted_amount),
            ", ".join(cat.tags) if cat.tags else "",
        ])


async def _sheet_currencies(wb, db: AsyncSession, user_id: int) -> None:
    ws = wb.create_sheet("Currencies")
    _write_header(ws, ["Code", "Symbol", "Name"])

    result = await db.execute(
        select(Currency)
        .where(Currency.user_id == user_id)
        .order_by(Currency.code)
    )
    for cur in result.scalars():
        ws.append([cur.code, cur.symbol, cur.name or ""])
